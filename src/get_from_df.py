'''
Project: Credit Ratings Prediction Using Neural Networks
Author: Antonio Pardo de Santayana Navarro
Description: Take .xlsx files generated with Factset Filings Wizard, and organize information into a pandas DataFrame
'''

import openpyxl as xl
import pandas as pd
import json
import os
import sys
import re

NUM_BATCHES = 5

# Some companies report in millions, some in thousands. We ignore share count for now and get the multiplier for each company
def assign_multiplier(text):
    if 'millions' in text or 'million' in text:
        return 1e6
    if 'thousand' in text or 'thousands' in text:
        return 1e3
    return 1

def clean_value(value):
    try:
        return float(re.sub(r"[^\d\.,-]", "", value))
    except ValueError:
        return 0
    except TypeError:
        return 0

def find_metric(ws, fields, multiplier):
    for row in ws.iter_rows(values_only=True):
        if row[0] and any(substring in row[0].lower() for substring in fields):
            value = row[1]
            if row[1] is None or 'other' in row[0].lower():
                continue
            if row[0].lower() == 'basic' and clean_value(row[1]) > 1000:
                continue
            if isinstance(value, str):
                value = clean_value(value)
            try:
                return value * multiplier if value else None
            except TypeError as e:
                print(e)
                sys.exit(1)
    return None

def process_file(file_path, patterns):
    print(f'Processing File {file_path}')
    wb = xl.load_workbook(file_path, data_only=True)
    companies = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        company={'Name': ws['A2'].value, 'Ticker': ws['A3'].value}
        multiplier = assign_multiplier(ws['A9'].value.lower())
        for name, fields in patterns.items():
            if name == 'EPS':
                company[name] = find_metric(ws, fields, 1)
            else:
                company[name] = find_metric(ws, fields, multiplier)
                if name == 'Liabilities and Equity':
                    try:
                        company['Total Liabilities'] = company['Liabilities and Equity'] - company['Equity']
                        del company['Liabilities and Equity']
                    except:
                        pass
        companies.append(company)
    with open('test.json', 'a') as f:
        json.dump(companies, f, indent=2)
    return companies


def main():
    full_data = []
    print("Last modified time:", os.path.getmtime('patterns.json'))
    with open('patterns.json', 'r') as f:
        patterns = json.load(f)
    for i in range(NUM_BATCHES):
        file_path = f'data/batch_{i}.xlsx'
        companies = process_file(file_path, patterns)
        full_data.extend(companies)
    df = pd.DataFrame(full_data)
    print(df.head())
    df.to_excel('full_data.xlsx', index=False)
        

if __name__ == '__main__':
    main()